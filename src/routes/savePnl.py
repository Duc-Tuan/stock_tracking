import MetaTrader5 as mt5
import os
import time
from datetime import datetime, timedelta, time as dtime
from src.models.modelSwapMt5 import SwapMt5
import multiprocessing
import pandas as pd
import json
from src.models.modelMultiAccountPnL import MultiAccountPnL
from src.models.model import SessionLocal
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from src.models.modelAccMt5 import AccountMt5
from src.utils.stop import stopDef
from filelock import FileLock
from openpyxl.utils.exceptions import InvalidFileException
import zipfile
from collections import defaultdict

def swap_difference(db, account_info):
    def get_log_for_day(day_date):
        start = datetime.combine(day_date, dtime(0, 0))              # 00:00:00
        end = datetime.combine(day_date, dtime(23, 59, 59, 999999))  # 23:59:59.999999
        return db.query(SwapMt5).filter(
            SwapMt5.username == account_info.login,
            SwapMt5.created_at.between(start, end)
        ).order_by(SwapMt5.created_at.desc()).first()

    now = datetime.now()
    today_date = now.date()
    yesterday_date = today_date - timedelta(days=1)

    # ✅ Tính mốc thời gian cắt ngày theo mùa
    if 4 <= now.month <= 9:
        cut_off = dtime(4, 3)  # mùa hè: 04:03
    else:
        cut_off = dtime(5, 3)  # mùa đông: 05:03

    if now.time() < cut_off:
        today_date = yesterday_date
        yesterday_date = today_date - timedelta(days=1)

    today_log = get_log_for_day(today_date)
    yesterday_log = get_log_for_day(yesterday_date)

    if yesterday_log and today_log:
        print(f"Tài khoản {account_info.login} swap hôm qua: {yesterday_log.swap}, hôm nay: {today_log.swap}")
        return today_log.swap - yesterday_log.swap

    return 0

def monitor_account(mt5_path, account_name, interval, queue, stop_event):
    if not mt5.initialize(path=mt5_path):
        print(f"[{account_name}] ❌ Cannot initialize MT5 at {mt5_path}")
        time.sleep(interval)
        return
    try: 
        while not stop_event.is_set():
            if stopDef(datetime.now()):
                print("⏸️ StopDef active: Dừng ghi log và theo dõi PnL vào thời điểm hiện tại")
                time.sleep(60)
                continue

            db = SessionLocal()
            
            try:
                account_info = mt5.account_info()
                positions = mt5.positions_get()

                # symbol_pnls = defaultdict(float)
                symbol_pnls_cvs_xlsx = {}
                symbol_pnls = {}
                for pos in positions:
                    symbol = pos.symbol
                    # Lấy giá thị trường hiện tại (tick)
                    tick = mt5.symbol_info_tick(symbol)
                    if tick:
                        # Có thể chọn giá BID hoặc ASK tùy theo loại lệnh:
                        # - BUY → tính theo BID (giá bạn có thể bán ra để đóng)
                        # - SELL → tính theo ASK (giá bạn phải mua lại để đóng)
                        current_price = tick.bid if pos.type == 0 else tick.ask
                        symbol_pnls_cvs_xlsx[symbol] = symbol_pnls_cvs_xlsx.get(symbol, 0.0) + current_price
                        # Nếu symbol chưa có, khởi tạo dict
                        if symbol not in symbol_pnls:
                            symbol_pnls[symbol] = {
                                "current_price": 0.0,
                                "type": "BUY" if pos.type == 0 else "SELL"
                            }
                        # Cộng dồn giá
                        symbol_pnls[symbol]["current_price"] += current_price

                existing = db.query(AccountMt5).filter(AccountMt5.username == account_info.login).all()
                if (len(existing) == 0):
                    new_data = AccountMt5(username=account_info.login, password='', loginId=1, server=account_info.server, by_symbol=json.dumps(list(symbol_pnls.keys())))
                    db.add(new_data)
                    db.commit()

                total_swap_difference = swap_difference(db, account_info)
                total_pnl = account_info.profit + abs(total_swap_difference)

                by_symbol_json_csv_file = json.dumps({k: round(v, 6) for k, v in symbol_pnls_cvs_xlsx.items()})
                by_symbol_json = json.dumps(symbol_pnls)
                num_positions = len(positions) if positions else 0

                if account_info:
                    data = {
                        "login": account_info.login,
                        "time": datetime.now().isoformat(),
                        "total_pnl": total_pnl,
                        "by_symbol": by_symbol_json_csv_file,
                        "num_positions": num_positions
                    }
                    log = MultiAccountPnL(
                        login=account_info.login,
                        total_pnl=total_pnl,
                        num_positions=num_positions,
                        time=datetime.now(),
                        by_symbol=by_symbol_json
                    )

                    db.add(log)
                    db.commit()
                    queue.put(data)  # 👉 gửi về process ghi log
                    print(f"✅ Đã ghi PnL {account_info.login}: giá chưa tính swap {account_info.profit}, giá đã tính swap {total_pnl} với {num_positions} lệnh, swap chênh lệch: {total_swap_difference}", f'info: {symbol_pnls}')
            except Exception as e:
                db.rollback()
                print(f"[{account_name}] ❌ Lỗi trong monitor_account: {e}")
            finally:
                time.sleep(interval)
    except KeyboardInterrupt:
        print("🔝 Logger process interrupted with Ctrl+C. Exiting gracefully.")
    finally:
        mt5.shutdown()


def logger_process(queue: multiprocessing.Queue, stop_event: multiprocessing.Event, csv_path="src/pnl_cache/pnl_log.csv", excel_path="src/pnl_cache/pnl_log.xlsx"): # type: ignore
    os.makedirs("src/pnl_cache", exist_ok=True)
    lock = FileLock(excel_path + ".lock")

    try:
        while not stop_event.is_set():
            data = queue.get()
            try:
                df = pd.DataFrame([data])
                if not os.path.exists(csv_path):
                    df.to_csv(csv_path, index=False)
                else:
                    df.to_csv(csv_path, mode="a", header=False, index=False)
            except Exception as e:
                print(f"⚠️ Error writing CSV: {e}")

            try:
                with lock:
                    if not os.path.exists(excel_path):
                        wb = Workbook()
                        ws = wb.active
                        ws.append(list(data.keys()))
                    else:
                        try:
                            wb = load_workbook(excel_path)
                            ws = wb.active
                        except (InvalidFileException, KeyError, zipfile.BadZipFile) as e:
                            print(f"⚠️ Excel file corrupted: {e} — Attempting to remove and recreate.")
                            try:
                                os.remove(excel_path)
                                print("🗑️ Corrupted Excel file removed.")
                            except PermissionError as remove_err:
                                print(f"❌ Cannot remove corrupted file: {remove_err}")
                                print("⚠️ Please close 'pnl_log.xlsx' if it is open in Excel.")
                                print("⏩ Skipping this write.")
                                continue
                            wb = Workbook()
                            ws = wb.active
                            ws.append(list(data.keys()))

                    ws.append(list(data.values()))
                    for i, column in enumerate(data.keys(), 1):
                        col_letter = get_column_letter(i)
                        max_length = max([len(str(cell.value)) for cell in ws[col_letter]] + [len(column)])
                        ws.column_dimensions[col_letter].width = max_length + 2

                    tmp_path = excel_path + ".tmp"
                    wb.save(tmp_path)
                    wb.close()  # ✅ Đảm bảo luôn close workbook!
                    os.replace(tmp_path, excel_path)

            except Exception as e:
                print(f"⚠️ Error writing Excel: {e}")

    except KeyboardInterrupt:
        print("🔝 Logger process interrupted with Ctrl+C. Exiting gracefully.")