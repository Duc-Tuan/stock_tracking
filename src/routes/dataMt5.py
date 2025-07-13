import MetaTrader5 as mt5
from src.models.model import SessionLocal
from datetime import datetime
from src.models.modelMultiAccountPnL import MultiAccountPnL
from src.models.modelAccMt5 import AccountMt5
from fastapi.encoders import jsonable_encoder
from src.middlewares.authMiddleware import decrypt_password_mt5
from src.controls.daily_email_sender import send_email_with_attachment
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import pandas as pd
import json
import asyncio
import csv
import os

async def log_one_account(acc):
    db = SessionLocal()
    login = acc["username"]

    try:
        ecrypted = decrypt_password_mt5(acc["password"])

        print(f"📥 Đang kết nối tài khoản {login}...")

        if not mt5.initialize(login=acc["username"], password=ecrypted, server=acc["server"]):
            print(f"❌ Login thất bại: {login}")
            return

        nameSymbol = []
        symbol_pnls = {}
        positions = mt5.positions_get()
        total_pnl = sum(pos.profit for pos in positions) if positions else 0.0
        num_positions = len(positions) if positions else 0

        for pos in positions:
            symbol = pos.symbol
            symbol_pnls[symbol] = symbol_pnls.get(symbol, 0.0) + pos.profit
            nameSymbol.append(pos.symbol)

        log = MultiAccountPnL(
            login=login,
            total_pnl=round(total_pnl, 2),
            num_positions=num_positions,
            time=datetime.now(),
            by_symbol=json.dumps({k: round(v, 2) for k, v in symbol_pnls.items()})
        )
        db.add(log)
        db.commit()

        # 👉 Ghi log ra CSV
        save_to_csv({
            "login": login,
            "total_pnl": round(total_pnl, 2),
            "num_positions": num_positions,
            "time": datetime.now().isoformat(),
            "by_symbol": json.dumps({k: round(v, 2) for k, v in symbol_pnls.items()})
        })

        # 👉 Ghi log ra Excel
        save_to_excel({
            "login": login,
            "total_pnl": round(total_pnl, 2),
            "num_positions": num_positions,
            "time": datetime.now().isoformat(),
            "by_symbol": json.dumps({k: round(v, 2) for k, v in symbol_pnls.items()})
        })

        print(f"✅ Đã ghi PnL {login}: {total_pnl} với {num_positions} lệnh {nameSymbol}")
    except Exception as e:
        print(f"❌ Lỗi tài khoản {login}:", e)
    finally:
        db.close()
        mt5.shutdown()

async def log_all_accounts_parallel():
    db = SessionLocal()
    account = db.query(AccountMt5).all()
    send_email_with_attachment()
    while True:
        print("🚀 Bắt đầu ghi log tất cả tài khoản song song...")
        tasks = [log_one_account(acc) for acc in jsonable_encoder(account)]
        await asyncio.gather(*tasks)
        await asyncio.sleep(1)  # chạy mỗi 10 giây

def save_to_csv(data: dict, filename="multi_account_pnl_log.csv"):
    file_exists = os.path.isfile(filename)
    
    with open(filename, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=[
            'login', 'total_pnl', 'num_positions', 'time', 'by_symbol'
        ])
        
        # Ghi header nếu file chưa tồn tại
        if not file_exists:
            writer.writeheader()
        
        writer.writerow(data)


# Hàm lưu dữ liệu vào file Excel
def save_to_excel(data: dict, filename="multi_account_pnl_log.xlsx"):
    # df_new = pd.DataFrame([data])  # Một dòng mới

    # if os.path.exists(filename):
    #     try:
    #         # 👉 Đọc file Excel cũ nếu có
    #         df_old = pd.read_excel(filename, engine='openpyxl')
    #         df_all = pd.concat([df_old, df_new], ignore_index=True)
    #     except Exception as e:
    #         print("❌ Lỗi khi đọc file Excel:", e)
    #         df_all = df_new
    # else:
    #     df_all = df_new

    # # 👉 Ghi dữ liệu vào file Excel
    
    # df_all.to_excel(filename, index=False, engine='openpyxl')
    if not os.path.exists(filename):
        # 👉 Tạo file mới nếu chưa có
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))  # Ghi header
    else:
        # 👉 Mở file có sẵn
        wb = load_workbook(filename)
        ws = wb.active

    # 👉 Ghi dữ liệu mới vào dòng tiếp theo
    ws.append(list(data.values()))

    # 👉 Tự động điều chỉnh độ rộng cột (tuỳ chọn)
    for i, column in enumerate(data.keys(), 1):
        col_letter = get_column_letter(i)
        max_length = max(
            [len(str(cell.value)) for cell in ws[col_letter]] + [len(column)]
        )
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)