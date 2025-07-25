import json
import os
import pandas as pd

from filelock import FileLock
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from datetime import datetime
from fastapi.responses import FileResponse
from openpyxl.utils import get_column_letter

from src.controls.authControll import get_current_user

router = APIRouter()

EXCEL_SOURCE = "src/pnl_cache/pnl_log.xlsx"

# 📍 Thư mục lưu file tạm thời
EXPORT_DIR = "exported_excels"
os.makedirs(EXPORT_DIR, exist_ok=True)

rename_map = {
    "login": "Tài khoản",
    "time": "Thời gian",
    "total_pnl": "Lợi nhuận tổng",
    "num_positions": "Số lệnh",
    "by_symbol": "Chi tiết theo cặp"
}

def delete_file(path: str):
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"🗑️ Đã xóa file tạm: {path}")
        except Exception as e:
            print(f"⚠️ Không thể xóa file {path}: {e}")

# ✅ Xử lý cột by_symbol an toàn
def parse_symbol(s):
    try:
        if s is None or s != s:
            return {}
        if isinstance(s, dict):
            return s
        if isinstance(s, str):
            s = s.strip()
            if s.lower() in ["", "null", "none"]:
                return {}
            return json.loads(s.replace("'", '"'))
        return {}
    except:
        return {}
    
@router.get("/download/pnl-log")
def download_excel(background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    if str(current_user.role) != "UserRole.admin":
        raise HTTPException(status_code=403, detail="Bạn không có quyền tải file")
    
    # 🔒 LOCK FILE TRƯỚC KHI ĐỌC
    lock = FileLock(EXCEL_SOURCE + ".lock")

    try:
        with lock:
            df = pd.read_excel(EXCEL_SOURCE)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi đọc file Excel gốc: {e}")

    df["by_symbol"] = df["by_symbol"].apply(parse_symbol)

    # 📁 Tạo tên file tạm và đường dẫn đầy đủ
    file_name = f"pnl_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(EXPORT_DIR, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for login_id, group in df.groupby("login"):
            group = group.reset_index(drop=True)

            # 👉 Chuyển mỗi dict trong 'by_symbol' thành cột mới cùng hàng
            expanded = group["by_symbol"].apply(pd.Series)
            merged_df = pd.concat([group.drop(columns=["by_symbol"]), expanded], axis=1)

            # 👉 Đổi tên cột nếu khớp với rename_map
            renamed_df = merged_df.rename(columns=rename_map)

            sheet_name = str(login_id)
            renamed_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 📏 Tự động điều chỉnh độ rộng cột
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(renamed_df.columns, 1):
                try:
                    max_len = renamed_df[col].apply(lambda x: len(str(x)) if pd.notna(x) else 0).max()
                    adjusted_width = min(max(max_len, len(str(col))) + 2, 50)
                    worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                except Exception as e:
                    print(f"⚠️ Lỗi khi xử lý cột {col}: {e}")
                    worksheet.column_dimensions[get_column_letter(idx)].width = 20

    # 👉 Đăng ký xóa file sau khi response trả về
    background_tasks.add_task(delete_file, file_path)

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file_name
    )


