import smtplib
import schedule
import os
import json
import time

import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from src.utils.options import SENDER_PASSWORD, SENDER_EMAIL, SEND_TIME

# ======== CẤU HÌNH NGƯỜI GỬI & NGƯỜI NHẬN =========
RECEIVER_EMAIL = ["testsendpymt5@gmail.com", "thaisanchezvn@gmail.com"]  # Hoặc danh sách: ["a@a.com", "b@b.com"]
ATTACHMENT_PATH = "src/pnl_cache/pnl_log.xlsx"  # Đường dẫn file đính kèm

rename_map = {
    "login": "Tài khoản",
    "time": "Thời gian",
    "total_pnl": "Lợi nhuận tổng",
    "num_positions": "Số lệnh",
    "by_symbol": "Chi tiết theo cặp"
}

# ✅ Xử lý cột by_symbol an toàn
def parse_symbol(s):
    try:
        if s is None or s != s:
            return {}
        if isinstance(s, dict):
            return s
        if isinstance(s, str):
            s = s.strip()
            if s.lower() in ["", "null", "none"]:
                return {}
            return json.loads(s.replace("'", '"'))
        return {}
    except:
        return {}

def send_email_with_attachment():
    print("📤 Đang gửi email...")

    # === Chuẩn bị email ===
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECEIVER_EMAIL if isinstance(RECEIVER_EMAIL, str) else ", ".join(RECEIVER_EMAIL)
    msg['Subject'] = "📊 Báo cáo định kỳ theo ngày kèm file"

    now = datetime.now()
    body = f"""Chào sếp,

Đây là báo cáo định kỳ vào lúc {SEND_TIME} ngày {now.strftime("%d/%m/%Y")} có kèm file đính kèm ở bên dưới.

Trân trọng."""
    msg.attach(MIMEText(body, 'plain'))

    # === Đọc dữ liệu gốc ===
    try:
        df = pd.read_excel(ATTACHMENT_PATH)
    except Exception as e:
        raise RuntimeError(f"Lỗi đọc file Excel gốc: {e}")

    # === Parse JSON column ===
    df["by_symbol"] = df["by_symbol"].apply(parse_symbol)

    # === Thư mục chứa file Excel từng login ===
    output_folder = "exported_excels"
    os.makedirs(output_folder, exist_ok=True)
    generated_files = []
    generated_files_delete = []

    # === Ghi từng file Excel ===
    for login_id, group in df.groupby("login"):
        group = group.reset_index(drop=True)
        if group.empty:
            print(f"⚠️ Bỏ qua login {login_id} vì không có dữ liệu.")
            continue

        try:
            expanded = group["by_symbol"].apply(pd.Series)
        except Exception as e:
            print(f"⚠️ Không thể phân tách 'by_symbol' cho login {login_id}: {e}")
            continue

        merged_df = pd.concat([group.drop(columns=["by_symbol"]), expanded], axis=1)
        renamed_df = merged_df.rename(columns=rename_map)
        if renamed_df.empty:
            print(f"⚠️ Dữ liệu sau xử lý cho login {login_id} rỗng. Bỏ qua.")
            continue

        file_name = os.path.join(output_folder, f"{login_id}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx")
        try:
            with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
                sheet_name = "data"
                renamed_df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(renamed_df.columns, 1):
                    try:
                        max_len = renamed_df[col].apply(lambda x: len(str(x)) if pd.notna(x) else 0).max()
                        adjusted_width = min(max(max_len, len(str(col))) + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        print(f"⚠️ Lỗi cột '{col}': {e}")
                        worksheet.column_dimensions[get_column_letter(idx)].width = 20

            print(f"✅ Đã ghi file: {file_name}")
            generated_files.append(file_name)
            generated_files_delete.append(file_name)
        except Exception as e:
            print(f"❌ Lỗi khi ghi file {file_name}: {e}")

    # === Ghi file CSV tổng (toàn bộ login) === 
    csv_file = os.path.join("src/pnl_cache", "pnl_log.csv")
    df.to_csv(csv_file, index=False)
    generated_files.append(csv_file)

    # === Đính kèm tất cả file vào email ===
    for file_path in generated_files:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename={os.path.basename(file_path)}'
                )
                msg.attach(part)
        else:
            print(f"⚠️ File không tồn tại: {file_path}")

    # === Gửi email ===
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
        server.quit()
        print("✅ Email đã được gửi thành công.")
    except Exception as e:
        print("❌ Lỗi khi gửi email:", e)

    # === (Tùy chọn) Xoá các file .xlsx sinh ra ===
    for file_path in generated_files_delete:
        if file_path.endswith(".xlsx"):
            try:
                os.remove(file_path)
                print(f"🧹 Đã xoá file: {file_path}")
            except Exception as e:
                print(f"⚠️ Không thể xoá file {file_path}: {e}")

# ========= LỊCH GỬI =========
schedule.every().day.at(SEND_TIME).do(send_email_with_attachment)

async def run_schedule_email():
    print(f"🕒 Script chạy, chờ gửi email mỗi ngày lúc {SEND_TIME}...")
    while True:
        schedule.run_pending()
        time.sleep(60)
